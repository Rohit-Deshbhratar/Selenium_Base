import csv
import inspect
import logging
from openpyxl import load_workbook


class Utils:
    def __init__(self):
        pass

    def read_data_from_excel(file_name, sheet):
        data_list = []
        wb = load_workbook(filename=file_name)
        sh = wb[sheet]
        row_count = sh.max_row
        column_count = sh.max_column

        for i in range(2, row_count + 1):
            row = []
            for j in range(1, column_count + 1):
                row.append(sh.cell(row=i, column=j).value)
            data_list.append(row)
        return data_list

    def read_data_from_csv(filename):
        data_list = []

        csv_data = open(filename, "r")
        reader = csv.reader(csv_data)
        next(reader)

        for rows in reader:
            data_list.append(rows)
        return data_list

    def custom_logger(log_level=logging.DEBUG):
        # set class/method name from where it is called
        logger_name = inspect.stack()[1][3]
        # create logger
        logger = logging.getLogger(logger_name)
        logger.setLevel(log_level)
        # create file handler and set log level
        fh = logging.FileHandler("selenium_base.log")
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(name)s : %(message)s',
                                      datefmt='%m/%d/%Y %I:%M:%S %p')
        # add formatter to file handler
        fh.setFormatter(formatter)

        # add file handler to logger
        logger.addHandler(fh)
        return logger
